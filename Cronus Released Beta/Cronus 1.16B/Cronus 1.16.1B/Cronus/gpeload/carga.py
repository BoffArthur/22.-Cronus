#Carga.py serve como automação de carga de dados das planilhas de viabilidades para o GPE. Aí surge o nome GPEAutomation
#Deve ser executado atraves do arquivo .exe disponivel no mesmo diretório deste aquivo.
#Para lançar uma nova versão, o código a ser executado é o abaixo para a biblioteca PyInstaller
# c:/Users/arthur.boff/AppData/Local/Programs/Python/Python310/Scripts/pyinstaller -c -F --icon carga.ico carga.py
#
#
#PARTE 0: DEFINIÇÃO DAS VARIAVEIS, BIBLIOTECAS E INPUTS
from glob import glob
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from openpyxl import workbook, load_workbook
from openpyxl.worksheet import worksheet
from win32com import client
from itertools import cycle
from time import sleep

def SepararParte():
    print('')
    print('--------------------------------------')
    print('')

def LoadForGPE(projetoIN):

    caminhoworkbook = "\\\\2431FSCXS01\\eng_produtoeprocesso$\\Viabilidades\\Ferramentas de Apoio\\Automação GPE\\"
    caminhoedgedriver = "\\\\2431FSCXS01\\eng_produtoeprocesso$\\Viabilidades\\Ferramentas de Apoio\\Automação GPE\\msedgedriver.exe"
    projetoIN = str(projetoIN)

    wb=load_workbook(filename= caminhoworkbook + 'PJT ' + projetoIN + '.xlsx', data_only=True) #carrega workbook
    AbaProjeto = "PJT " + projetoIN #faz a junção das strings "PJT" e o número do projeto que o usuário entrou
    ws=wb[AbaProjeto] #carrega worksheet [NOME] do workbook

    #Definição das variaveis

    projeto = ws.cell(1, 2).value #'Projeto N°'
    print('Dados encontrados do Projeto N°  ' + str(projeto)) 

    larguratira = ws.cell(2, 2).value #'Lagura da Tira [mm]'
    print('Lagura da Tira [mm]  '+str(larguratira)) 

    maquina1 = ws.cell(3, 2).value #'Máquina'
    print('Máquina 1  '+str(maquina1))

    ajustemetros = ws.cell(4, 2).value #'Ajuste [m]'
    if ajustemetros==None: ajustemetros=0
    print('Ajuste [m]  '+str(ajustemetros))

    ajustehoras = ws.cell(5, 2).value #'Ajuste [horas]'
    if ajustehoras==None: ajustehoras= 0
    print('Ajuste [horas]  '+str(ajustehoras))

    velocidade = ws.cell(6, 2).value #'Velocidade [m/min]'
    print('Velocidade [m/min]  '+str(velocidade))

    setuptotal = ws.cell(7, 2).value #'Setup Total [horas]'
    print('Setup Total [horas]  '+str(setuptotal))

    perdasPP = ws.cell(8, 2).value #'Perdas Pré Punching [%]'
    print('Perdas Pré Punching [%]  '+str(perdasPP))

    perdastotais = ws.cell(9, 2).value #'Perdas Totais [%]'
    print('Perdas Totais [%]  '+str(perdastotais))

    eficiencia = ws.cell(10, 2).value #'Eficiência [%]'
    print('Eficiência [%]  '+str(eficiencia))

    opSlliter = ws.cell(11, 2).value #'Operação Slliter [min/ton]'
    print('Operação Slliter [min/ton]  '+str(opSlliter))

    prSlliter = ws.cell(12, 2).value #'Preparação Slliter [min/ton]'
    print('Preparação Slliter [min/ton]  '+str(prSlliter))

    perdasSlliter = ws.cell(13, 2).value #'Perdas Slliter [%]'
    print('Perdas Slliter [%]  '+str(perdasSlliter))

    Investimento = ws.cell(14, 2).value #'Total INVESTIMENTO: [R$]'
    if Investimento==None: Investimento=0
    print('Total INVESTIMENTO: [R$]  '+str(Investimento))

    maquina1SN = ws.cell(15, 2).value #'Máquina1 ShortName'
    print('Máquina 1 ShortName: '+str(maquina1SN))

    SetupOff = ws.cell(16, 2).value #Tempo de Setup Off-line [h]
    if SetupOff==None: SetupOff=0
    print('Tempo Setup OffLine [h]: '+str(SetupOff))

    BitolaOrg = ws.cell(17, 2).value #Tempo de Setup Off-line [h]
    if BitolaOrg==None: BitolaOrg=0
    print('Bitola de Origem [mm]: '+str(BitolaOrg))

    SepararParte()

    ProjetoGPE = projetoIN[0:4] # [0:4] pega apenas os caracteres de 0 a 4 na string
    maquina1GPE = maquina1[:-2].lower() #[:-2] retira os dois ultimos caracteres ; lower() deixa todos os caracteres minusculos

    print("Gestão de Projetos de Engenharia / Carregando...")
    driver = webdriver.Edge(caminhoedgedriver)
    driver.set_window_size(1567, 423)
    driver.minimize_window() #minimizar janela
    driver.get("http://intranetmeincol/intranet/gpe/relatorio_acompanhamento.php") #ir para URL definida
    #driver.maximize_window() #maximizar janela

    driver.find_element(By.ID, "usuario").send_keys("eng_produto6") #entrar usuario
    driver.find_element(By.ID, "senha").send_keys("vMeincol\\20191") #entrar senha
    driver.find_element(By.CSS_SELECTOR, "td:nth-child(1) > input").click()

    print("Projeto no GPE: "+ProjetoGPE+" / Carregando...")
    driver.find_element(By.ID, "buscar").send_keys(ProjetoGPE) #entrar numero do projeto
    driver.find_element(By.ID, "comboBusca").click() #clicar em buscar
    dropdown = driver.find_element(By.ID, "comboBusca")

    print(projetoIN +" / 2.3 Viabilidade Industrial / Carregando...")
    dropdown.find_element(By.XPATH, "//option[. = 'Projeto']").click()
    driver.find_element(By.CSS_SELECTOR, ".btnBusca > input").click()
    driver.find_element(By.LINK_TEXT, projetoIN).click() #buscar a subpasta do projeto (subprojeto ou partição)
    driver.find_element(By.LINK_TEXT, "2.3 - Viabilidade Industrial").click()

    #caracteristicas do produto
    print('     Preenchendo Características do Produto:')
    print('         Bitola de Origem: '+ str(BitolaOrg))
    print('         Largura da Tira: '+ str(larguratira))
    print('         Perdas Scrap: '+ str(perdasPP)[0:4].replace(".",","))
    driver.find_element(By.ID, "matriz_origem_mm").send_keys(Keys.BACKSPACE*10, str(BitolaOrg).replace(".",",")) #inserir a bitola de origem
    driver.find_element(By.ID, "largura_tira").send_keys(Keys.BACKSPACE*10, str(larguratira).replace(".",",")) #inserir a largura da tira
    driver.find_element(By.CSS_SELECTOR, ".tabela:nth-child(10) tr:nth-child(6)").click()
    driver.find_element(By.NAME, "perfuracao").send_keys(Keys.BACKSPACE*10, str(perdasPP)[0:4].replace(".",",")) #inserir a perda por perfuração SCRAP
    driver.find_element(By.ID, "salvar").click() #Salvar

    #operações
    print('     Preenchendo operações:')
    print('         Operação 1: Slliter')
    print('         Operação 2: '+maquina1GPE)
    print('         Operação 3: ')

    if driver.find_element(By.ID, "operacao_slitter").is_selected() == False: driver.find_element(By.ID, "operacao_slitter").click() #somente selecionar se estiver desmarcado
    if driver.find_element(By.ID, "operacao_"+ maquina1GPE ).is_selected() == False: 
        driver.find_element(By.ID, "operacao_"+ maquina1GPE ).click() #somente selecionar se estiver desmarcado
    if maquina1GPE == "formadora" and driver.find_element(By.ID, "operacao_perfiladeira").is_selected()==True: driver.find_element(By.ID, "operacao_perfiladeira").click()
    if maquina1GPE == "perfiladeira" and driver.find_element(By.ID, "operacao_formadora").is_selected()==True: driver.find_element(By.ID, "operacao_formadora").click()
    dropdown = driver.find_element(By.ID, "ano_fiscal") #define variavel para ano fiscal   
    dropdown.find_element(By.XPATH, "//option[. = '2023']").click() #define o ano fiscal corrente
    driver.find_element(By.ID, "salvar").click() #Salvar
    #driver.find_element(By.LINK_TEXT, "2.3 - Viabilidade Industrial").click() #atualiza a pagina clicando novamente em viabilidade industrial

    driver.find_element(By.ID, "oper1_maquina").click()
    dropdown = driver.find_element(By.ID, "oper1_maquina") #seleciona a operação
    dropdown.find_element(By.XPATH, "//option[. = 'S02 ']").click() #seleciona a maquina da operação 1

    driver.find_element(By.ID, "oper2_maquina").click()
    dropdown = driver.find_element(By.ID, "oper2_maquina") 
    dropdown.find_element(By.XPATH, "//option[. = '" + maquina1SN + "']").click() #preenche a máquina da operação 2
    driver.find_element(By.ID, "obs").clear() #limpar todos os dados das observações
    driver.find_element(By.ID, "obs").send_keys("REV00: Análise técnica nos anexos.") #preencher as observações com a nota padrão

    #Preencher checklist
    if maquina1GPE == 'perfiladeira': 
        print('     Preenchendo checklist: Padrão Perfiladeira')
        driver.find_element(By.CSS_SELECTOR, "tr:nth-child(4) tr:nth-child(2) > td:nth-child(3)").click()
        driver.find_element(By.NAME, "status2").click()
        driver.find_element(By.CSS_SELECTOR, "tr:nth-child(4) > td:nth-child(5) > input").click()
        driver.find_element(By.NAME, "status4").click()
        driver.find_element(By.NAME, "status5").click()
        driver.find_element(By.NAME, "status6").click()
        driver.find_element(By.NAME, "status7").click()
        driver.find_element(By.NAME, "status8").click()
        driver.find_element(By.NAME, "status9").click()
        driver.find_element(By.NAME, "status10").click()
        driver.find_element(By.NAME, "status11").click()
        driver.find_element(By.NAME, "status12").click()
        driver.find_element(By.CSS_SELECTOR, "tr:nth-child(14) > td:nth-child(5) > input").click()
        driver.find_element(By.CSS_SELECTOR, "tr:nth-child(15) > td:nth-child(5) > input").click()
        driver.find_element(By.NAME, "status15").click()
        driver.find_element(By.NAME, "status16").click()

    if maquina1GPE == 'formadora':
        print('     Preenchendo checklist: Padrão Formadora')
        driver.find_element(By.NAME, "status1").click()
        driver.find_element(By.NAME, "status2").click()
        driver.find_element(By.NAME, "status3").click()
        driver.find_element(By.NAME, "status4").click()
        driver.find_element(By.NAME, "status5").click()
        driver.find_element(By.NAME, "status6").click()
        driver.find_element(By.NAME, "status7").click()
        driver.find_element(By.NAME, "status8").click()
        driver.find_element(By.NAME, "status9").click()
        driver.find_element(By.NAME, "status10").click()
        driver.find_element(By.NAME, "status11").click()
        driver.find_element(By.NAME, "status12").click()
        driver.find_element(By.CSS_SELECTOR, "tr:nth-child(14) > td:nth-child(5) > input").click()
        driver.find_element(By.CSS_SELECTOR, "tr:nth-child(15) > td:nth-child(5) > input").click()
        driver.find_element(By.NAME, "status15").click()
        driver.find_element(By.NAME, "status16").click()

    #investimentos e estrutura de fluxo de processo
    print('2.3.1 Preenchendo investimentos e estrutura de fluxo de processo ...')
    driver.find_element(By.NAME, "preencher").click() #Entra na aba preencher para entrar os dados

    #DADOS DE OPERAÇÕES

    #operação 1
    driver.find_element(By.ID, "oper1_tempo_operacao").send_keys(Keys.BACKSPACE*10, str(opSlliter)[0:4].replace(".",",")) #Tempo Operação Slliter[h]
    driver.find_element(By.ID, "oper1_tempo_preparacao").send_keys(Keys.BACKSPACE*10, str(prSlliter)[0:4].replace(".",",")) #Tempo Preparação Slliter[h]
    driver.find_element(By.ID, "oper1_perdas").send_keys(Keys.BACKSPACE*10, str(perdasSlliter)[0:4].replace(".",",")) #Perdas Slliter [%]
    print('     Operação 1 ok...')

    #operação 2
    if maquina1GPE == 'perfiladeira':
        perdasPROD = float(str(driver.find_element(By.ID, "oper2_perdas_ajuste").get_attribute("value")).replace(',','.'))
        driver.find_element(By.ID, "oper2_utilizacao").send_keys(Keys.BACKSPACE*10, eficiencia) #eficiencia[%]
        driver.find_element(By.ID, "oper2_ajuste").send_keys(Keys.BACKSPACE*10, ajustemetros) #pedas de ajuste [m]
        driver.find_element(By.ID, "oper2_perdas_producao").send_keys(Keys.BACKSPACE*10, str(perdastotais - perdasPROD)[0:4].replace(".",",")) #perdas de produção [m] 
        driver.find_element(By.ID, "oper2_velocidade").send_keys(Keys.BACKSPACE*10, velocidade) #velocidade da linha [m/min]
        driver.find_element(By.ID, "oper2_ts_fer").send_keys(Keys.BACKSPACE*10, str(setuptotal - ajustehoras)[0:4].replace(".",",")) #tempo de setup do ferramental [h] 
        driver.find_element(By.ID, "oper2_ts_ajuste").send_keys(Keys.BACKSPACE*10, str(ajustehoras)[0:4].replace(".",",")) #tempo de ajuste
        driver.find_element(By.ID, "oper2_investimento_fer").send_keys(Keys.BACKSPACE*10, str(Investimento).replace(".",",")) #investimento [R$]

    if maquina1GPE == 'formadora':
        driver.find_element(By.ID, "oper2_utilizacao").send_keys(Keys.BACKSPACE*10, eficiencia) #eficiencia[%]
        driver.find_element(By.ID, "oper2_ajuste").send_keys(Keys.BACKSPACE*10, ajustemetros) #pedas de ajuste [m]
        driver.find_element(By.ID, "oper2_perdas_producao").send_keys(Keys.BACKSPACE*10, 0) #perdas de produção [m] [zera o valor para calcular corretamente]
        perdasPROD = float(str(driver.find_element(By.ID, "oper2_perdas_ajuste").get_attribute("value")).replace(',','.')) #faz o calculo das perdas de PRODUÇÃO
        driver.find_element(By.ID, "oper2_perdas_producao").send_keys(Keys.BACKSPACE*10, str(perdastotais - perdasPROD)[0:4].replace(".",",")) #perdas de produção [m] 
        driver.find_element(By.ID, "oper2_velocidade").send_keys(Keys.BACKSPACE*10, velocidade) #velocidade da linha [m/min]
        driver.find_element(By.ID, "oper2_ts_fer").send_keys(Keys.BACKSPACE*10, str(setuptotal - ajustehoras)[0:4].replace(".",",")) #tempo de setup do ferramental [h] 
        driver.find_element(By.ID, "oper2_ts_ajuste").send_keys(Keys.BACKSPACE*10, str(ajustehoras)[0:4].replace(".",",")) #tempo de ajuste
        driver.find_element(By.ID, "oper2_investimento_fer").send_keys(Keys.BACKSPACE*10, str(Investimento).replace(".",",")) #investimento [R$]
        try:
            driver.find_element(By.ID, "oper2_setupoff_h").send_keys(Keys.BACKSPACE*10, SetupOff) #Tempo setup OffLine [h]
        except:
            None

    print('     Operação 2 ok...')
    driver.find_element(By.ID, "salvar").click()
    driver.find_element(By.LINK_TEXT, "Viabilidade Industrial").click()
    SepararParte()

    print('Salvando e saindo...')
    driver.find_element(By.LINK_TEXT, "AE").click()
    driver.find_element(By.LINK_TEXT, "Sair").click()

    driver.close

    print("Leitura do xlsx finalizada. Carga ao GPE finalizada. Browser Fechado. Fim")
    print('')

#LoadForGPE(2538.1)