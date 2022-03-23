print('C.R.O.N.U.S.... Starting the application and decrompressing the required files...')
#Carga.py serve como automação de carga de dados das planilhas de viabilidades para o GPE. Aí surge o nome GPEAutomation
#Deve ser executado atraves do arquivo .exe disponivel no mesmo diretório deste aquivo.
#Para lançar uma nova versão, o código a ser executado é o abaixo para a biblioteca PyInstaller
# c:/Users/arthur.boff/AppData/Local/Programs/Python/Python310/Scripts/pyinstaller -c -F --icon Carga.ico carga.py
# c:/Users/arthur.boff/AppData/Local/Programs/Python/Python310/Scripts/pyinstaller -c -D --icon Carga.ico carga.py
#
#PARTE 0: DEFINIÇÃO DAS VARIAVEIS, BIBLIOTECAS E INPUTS
from cgitb import text
from ctypes.wintypes import MSG
from glob import glob
from pyexpat import model
from sqlite3 import Row
import string
from turtle import clone
from pyparsing import col
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from openpyxl import workbook, load_workbook
from openpyxl.worksheet import worksheet
from win32com import client
from numpy import row_stack, true_divide
from tkinter import HORIZONTAL, StringVar, ttk
from tkinter.messagebox import showinfo
from tkinter import messagebox
import tkinter as tk
from tkinter import ttk
from sys import exit
import os
import time 

#sobre um nível da pasta atual
AutomationDir = str(os.getcwd())
while AutomationDir[-1] != '\\':
    AutomationDir = AutomationDir[:-1]
AutomationDir = AutomationDir[:-1]
os.chdir(AutomationDir)

#slpah screen window
splash_win= tk.Tk()
splash_win.title("Cronus ™")
splash_win.geometry("700x200")
splash_win.eval('tk::PlaceWindow . center')
splash_win.overrideredirect(True)
splash_label= ttk.Label(splash_win, text= "C.R.O.N.U.S.", foreground= "red", font= ('BankGothic Lt BT', 40)).place(x=350, y=100, anchor="center")

# store strings
user = tk.StringVar()
password = tk.StringVar()
project = tk.StringVar()

def SepararParte():
    print('')
    print('--------------------------------------')
    print('')     

def load_clicked():
    #verifica se o usuário preencheu os dados corretamente
    if user.get()=="" or password.get()=="" or project.get()=="":
        p1.stop()
        MsgBox = tk.messagebox.showerror(title='Cronus', message='Dados incompletos. Favor preencher todos os campos solicitados.')
    else:
        p1.start()
        root_update()
        projetoIN=project.get()
        try:
            wb=load_workbook(filename= 'PJT ' + projetoIN + '.xlsx', data_only=True) #carrega workbook
            AbaProjeto = "PJT " + projetoIN #faz a junção das strings "PJT" e o número do projeto que o usuário entrou
            ws=wb[AbaProjeto] #carrega worksheet [NOME] do workbook

            #Definição das variaveis
            projeto = ws.cell(1, 2).value #'Projeto N°'
            larguratira = ws.cell(2, 2).value #'Lagura da Tira [mm]'
            maquina1 = ws.cell(3, 2).value #'Máquina'
            ajustemetros = ws.cell(4, 2).value #'Ajuste [m]'
            ajustehoras = ws.cell(5, 2).value #'Ajuste [horas]'
            velocidade = ws.cell(6, 2).value #'Velocidade [m/min]'
            setuptotal = ws.cell(7, 2).value #'Setup Total [horas]'
            perdasPP = ws.cell(8, 2).value #'Perdas Pré Punching [%]'
            perdastotais = ws.cell(9, 2).value #'Perdas Totais [%]'
            eficiencia = ws.cell(10, 2).value #'Eficiência [%]'
            opSlliter = ws.cell(11, 2).value #'Operação Slliter [min/ton]'
            prSlliter = ws.cell(12, 2).value #'Preparação Slliter [min/ton]'
            perdasSlliter = ws.cell(13, 2).value #'Perdas Slliter [%]'
            Investimento = ws.cell(14, 2).value #'Total INVESTIMENTO: [R$]'
            maquina1SN = ws.cell(15, 2).value #'Máquina1 ShortName'
            SetupOff = ws.cell(16, 2).value #Tempo de Setup Off-line [h]

            #caso alguma variável seja Nula definir como 0
            if ajustehoras==None: ajustehoras= 0
            if ajustemetros==None: ajustemetros=0
            if Investimento==None: Investimento=0
            if SetupOff==None: SetupOff=0

            #print das variaveis ao usuário
            projeto_label = ttk.Label(root, text='Dados encontrados do Projeto N°  ' + str(projeto)).grid(column=0, row=5, sticky=tk.W, columnspan=2)
            larguratira_label = ttk.Label(root, text='Lagura da Tira [mm]  '+str(larguratira)).grid(column=0, row=6, sticky=tk.W, columnspan=2)
            maquina1_label = ttk.Label(root, text='Máquina 1  '+str(maquina1)).grid(column=0, row=7, sticky=tk.W, columnspan=2)
            ajustemetros_label = ttk.Label(root, text='Ajuste [m]  '+str(ajustemetros)).grid(column=0, row=8, sticky=tk.W, columnspan=2)
            ajustehoras_label = ttk.Label(root, text='Ajuste [horas]  '+str(ajustehoras)).grid(column=0, row=9, sticky=tk.W, columnspan=2)
            velocidade_label = ttk.Label(root, text='Velocidade [m/min]  '+str(velocidade)).grid(column=0, row=10, sticky=tk.W, columnspan=2)
            setuptotal_label = ttk.Label(root, text='Setup Total [horas]  '+str(setuptotal)).grid(column=0, row=11, sticky=tk.W, columnspan=2)
            perdasPP_label = ttk.Label(root, text='Perdas Pré Punching [%]  '+str(perdasPP)).grid(column=0, row=12, sticky=tk.W, columnspan=2)
            perdastotais_label = ttk.Label(root, text='Perdas Totais [%]  '+str(perdastotais)).grid(column=0, row=13, sticky=tk.W, columnspan=2)
            eficiencia_label = ttk.Label(root, text='Eficiência [%]  '+str(eficiencia)).grid(column=0, row=14, sticky=tk.W, columnspan=2)
            opSlliter_label = ttk.Label(root, text='Operação Slliter [min/ton]  '+str(opSlliter)).grid(column=0, row=15, sticky=tk.W, columnspan=2)
            prSlliter_label = ttk.Label(root, text='Preparação Slliter [min/ton]  '+str(prSlliter)).grid(column=0, row=16, sticky=tk.W, columnspan=2)
            perdasSlliter_label = ttk.Label(root, text='Perdas Slliter [%]  '+str(perdasSlliter)).grid(column=0, row=17, sticky=tk.W, columnspan=2)
            Investimento_label = ttk.Label(root, text='Total INVESTIMENTO: [R$]  '+str(Investimento)).grid(column=0, row=18, sticky=tk.W, columnspan=2)
            maquina1SN_label = ttk.Label(root, text='Máquina 1 ShortName: '+str(maquina1SN)).grid(column=0, row=19, sticky=tk.W, columnspan=2)
            SetupOff_label = ttk.Label(root, text='Tempo Setup OffLine [h]: '+str(SetupOff)).grid(column=0, row=20, sticky=tk.W, columnspan=2)
            
            ProjetoGPE = projetoIN[0:4] # [0:4] pega apenas os caracteres de 0 a 4 na string
            maquina1GPE = maquina1[:-2].lower() #[:-2] retira os dois ultimos caracteres ; lower() deixa todos os caracteres minusculos

            status.set('Gestão de Projetos de Engenharia / Carregando...')
            root_update()
            driver = webdriver.Edge()
            driver.set_window_size(1567, 423)
            driver.minimize_window() #minimizar janela
            driver.get("http://intranetmeincol/intranet/gpe/relatorio_acompanhamento.php") #ir para URL definida

            driver.find_element(By.ID, "usuario").send_keys(user.get()) #entrar usuário
            driver.find_element(By.ID, "senha").send_keys(password.get()) #entrar senha 
            driver.find_element(By.ID, "usuario").send_keys(Keys.ENTER)

            try:
                driver.find_element(By.ID, "buscar").send_keys(ProjetoGPE)  #entrar numero do projeto
                status.set("Projeto no GPE: "+ProjetoGPE+" / Carregando...")
                root_update()
                driver.find_element(By.ID, "comboBusca").click() #clicar em buscar
                dropdown = driver.find_element(By.ID, "comboBusca")

                status.set(projetoIN +" / 2.3 Viabilidade Industrial / Carregando...")
                dropdown.find_element(By.XPATH, "//option[. = 'Projeto']").click()
                driver.find_element(By.CSS_SELECTOR, ".btnBusca > input").click()
                driver.find_element(By.LINK_TEXT, projetoIN).click() #buscar a subpasta do projeto (subprojeto ou partição)
                driver.find_element(By.LINK_TEXT, "2.3 - Viabilidade Industrial").click()

                #caracteristicas do produto
                status.set('Preenchendo Características do Produto...')
                root_update()
                driver.find_element(By.ID, "largura_tira").send_keys(Keys.BACKSPACE*10, str(larguratira).replace(".",",")) #inserir a largura da tira
                driver.find_element(By.CSS_SELECTOR, ".tabela:nth-child(10) tr:nth-child(6)").click()
                driver.find_element(By.NAME, "perfuracao").send_keys(Keys.BACKSPACE*10, str(perdasPP)[0:4].replace(".",",")) #inserir a perda por perfuração SCRAP
                driver.find_element(By.ID, "salvar").click() #Salvar

                #operações
                status.set('Preenchendo operações...')
                root_update()

                if driver.find_element(By.ID, "operacao_slitter").is_selected() == False: driver.find_element(By.ID, "operacao_slitter").click() #somente selecionar se estiver desmarcado
                if driver.find_element(By.ID, "operacao_"+ maquina1GPE ).is_selected() == False: 
                    driver.find_element(By.ID, "operacao_"+ maquina1GPE ).click() #somente selecionar se estiver desmarcado
                if maquina1GPE == "formadora" and driver.find_element(By.ID, "operacao_perfiladeira").is_selected()==True: driver.find_element(By.ID, "operacao_perfiladeira").click()
                if maquina1GPE == "perfiladeira" and driver.find_element(By.ID, "operacao_formadora").is_selected()==True: driver.find_element(By.ID, "operacao_formadora").click()
                dropdown = driver.find_element(By.ID, "ano_fiscal") #define variavel para ano fiscal   
                dropdown.find_element(By.XPATH, "//option[. = '2022']").click() #define o ano fiscal corrente
                driver.find_element(By.ID, "salvar").click() #Salvar
                #driver.find_element(By.LINK_TEXT, "2.3 - Viabilidade Industrial").click() #atualiza a pagina clicando novamente em viabilidade industrial
                
                driver.find_element(By.ID, "oper1_maquina").click()
                dropdown = driver.find_element(By.ID, "oper1_maquina") #seleciona a operação
                dropdown.find_element(By.XPATH, "//option[. = 'S02 ']").click() #seleciona a maquina da operação 1
                
                driver.find_element(By.ID, "oper2_maquina").click()
                dropdown = driver.find_element(By.ID, "oper2_maquina") 
                dropdown.find_element(By.XPATH, "//option[. = '" + maquina1SN + "']").click() #preenche a máquina da operação 2
                driver.find_element(By.ID, "obs").clear() #limpar todos os dados das observações
                driver.find_element(By.ID, "obs").send_keys("REV00: Análise técnica nos anexos.") #preencher as observações com a nota padrão
                
                #Preencher checklist
                if maquina1GPE == 'perfiladeira': 
                    status.set('Preenchendo checklist: Padrão Perfiladeira...')
                    root_update()
                    driver.find_element(By.CSS_SELECTOR, "tr:nth-child(4) tr:nth-child(2) > td:nth-child(3)").click()
                    driver.find_element(By.NAME, "status2").click()
                    driver.find_element(By.CSS_SELECTOR, "tr:nth-child(4) > td:nth-child(5) > input").click()
                    driver.find_element(By.NAME, "status4").click()
                    driver.find_element(By.NAME, "status5").click()
                    driver.find_element(By.NAME, "status6").click()
                    driver.find_element(By.NAME, "status7").click()
                    driver.find_element(By.NAME, "status8").click()
                    driver.find_element(By.NAME, "status9").click()
                    driver.find_element(By.NAME, "status10").click()
                    driver.find_element(By.NAME, "status11").click()
                    driver.find_element(By.NAME, "status12").click()
                    driver.find_element(By.CSS_SELECTOR, "tr:nth-child(14) > td:nth-child(5) > input").click()
                    driver.find_element(By.CSS_SELECTOR, "tr:nth-child(15) > td:nth-child(5) > input").click()
                    driver.find_element(By.NAME, "status15").click()
                    driver.find_element(By.NAME, "status16").click()
                
                if maquina1GPE == 'formadora':
                    status.set('Preenchendo checklist: Padrão Formadora...')
                    root_update()
                    driver.find_element(By.NAME, "status1").click()
                    driver.find_element(By.NAME, "status2").click()
                    driver.find_element(By.NAME, "status3").click()
                    driver.find_element(By.NAME, "status4").click()
                    driver.find_element(By.NAME, "status5").click()
                    driver.find_element(By.NAME, "status6").click()
                    driver.find_element(By.NAME, "status7").click()
                    driver.find_element(By.NAME, "status8").click()
                    driver.find_element(By.NAME, "status9").click()
                    driver.find_element(By.NAME, "status10").click()
                    driver.find_element(By.NAME, "status11").click()
                    driver.find_element(By.NAME, "status12").click()
                    driver.find_element(By.CSS_SELECTOR, "tr:nth-child(14) > td:nth-child(5) > input").click()
                    driver.find_element(By.CSS_SELECTOR, "tr:nth-child(15) > td:nth-child(5) > input").click()
                    driver.find_element(By.NAME, "status15").click()
                    driver.find_element(By.NAME, "status16").click()

                #investimentos e estrutura de fluxo de processo
                status.set('2.3.1 Preenchendo investimentos e estrutura de fluxo de processo...')
                root_update()
                driver.find_element(By.NAME, "preencher").click() #Entra na aba preencher para entrar os dados

                #DADOS DE OPERAÇÕES
                #operação 1
                driver.find_element(By.ID, "oper1_tempo_operacao").send_keys(Keys.BACKSPACE*10, str(opSlliter)[0:4].replace(".",",")) #Tempo Operação Slliter[h]
                driver.find_element(By.ID, "oper1_tempo_preparacao").send_keys(Keys.BACKSPACE*10, str(prSlliter)[0:4].replace(".",",")) #Tempo Preparação Slliter[h]
                driver.find_element(By.ID, "oper1_perdas").send_keys(Keys.BACKSPACE*10, str(perdasSlliter)[0:4].replace(".",",")) #Perdas Slliter [%]

                #operação 2
                if maquina1GPE == 'perfiladeira':
                    perdasPROD = float(str(driver.find_element(By.ID, "oper2_perdas_ajuste").get_attribute("value")).replace(',','.'))
                    driver.find_element(By.ID, "oper2_utilizacao").send_keys(Keys.BACKSPACE*10, eficiencia) #eficiencia[%]
                    driver.find_element(By.ID, "oper2_ajuste").send_keys(Keys.BACKSPACE*10, ajustemetros) #pedas de ajuste [m]
                    driver.find_element(By.ID, "oper2_perdas_producao").send_keys(Keys.BACKSPACE*10, str(perdastotais - perdasPROD)[0:4].replace(".",",")) #perdas de produção [m] 
                    driver.find_element(By.ID, "oper2_velocidade").send_keys(Keys.BACKSPACE*10, velocidade) #velocidade da linha [m/min]
                    driver.find_element(By.ID, "oper2_ts_fer").send_keys(Keys.BACKSPACE*10, str(setuptotal - ajustehoras)[0:4].replace(".",",")) #tempo de setup do ferramental [h] 
                    driver.find_element(By.ID, "oper2_ts_ajuste").send_keys(Keys.BACKSPACE*10, str(ajustehoras)[0:4].replace(".",",")) #tempo de ajuste
                    driver.find_element(By.ID, "oper2_investimento_fer").send_keys(Keys.BACKSPACE*10, Investimento) #investimento [R$]
                
                if maquina1GPE == 'formadora':
                    driver.find_element(By.ID, "oper2_utilizacao").send_keys(Keys.BACKSPACE*10, eficiencia) #eficiencia[%]
                    driver.find_element(By.ID, "oper2_ajuste").send_keys(Keys.BACKSPACE*10, ajustemetros) #pedas de ajuste [m]
                    driver.find_element(By.ID, "oper2_perdas_producao").send_keys(Keys.BACKSPACE*10, 0) #perdas de produção [m] [zera o valor para calcular corretamente]
                    perdasPROD = float(str(driver.find_element(By.ID, "oper2_perdas_ajuste").get_attribute("value")).replace(',','.')) #faz o calculo das perdas de PRODUÇÃO
                    driver.find_element(By.ID, "oper2_perdas_producao").send_keys(Keys.BACKSPACE*10, str(perdastotais - perdasPROD)[0:4].replace(".",",")) #perdas de produção [m] 
                    driver.find_element(By.ID, "oper2_velocidade").send_keys(Keys.BACKSPACE*10, velocidade) #velocidade da linha [m/min]
                    driver.find_element(By.ID, "oper2_ts_fer").send_keys(Keys.BACKSPACE*10, str(setuptotal - ajustehoras)[0:4].replace(".",",")) #tempo de setup do ferramental [h] 
                    driver.find_element(By.ID, "oper2_ts_ajuste").send_keys(Keys.BACKSPACE*10, str(ajustehoras)[0:4].replace(".",",")) #tempo de ajuste
                    driver.find_element(By.ID, "oper2_investimento_fer").send_keys(Keys.BACKSPACE*10, Investimento) #investimento [R$]
                    driver.find_element(By.ID, "oper2_setupoff_h").send_keys(Keys.BACKSPACE*10, SetupOff) #Tempo setup OffLine [h]

                driver.find_element(By.ID, "salvar").click()
                driver.find_element(By.LINK_TEXT, "Viabilidade Industrial").click()
                
                status.set('Salvando e saindo...')
                root_update()
                driver.find_element(By.LINK_TEXT, "AE").click()
                driver.find_element(By.LINK_TEXT, "Sair").click()
                
                driver.close
                status.set("Carga ao GPE do projeto " + str(project.get()) + " finalizada.")
                root_update()
            except NoSuchElementException:
                p1.stop()
                MsgBox = tk.messagebox.showerror(title='Cronus', message='Usuário ou Senha inválidos.')
        except FileNotFoundError:
            p1.stop()
            MsgBox = tk.messagebox.showerror(title='Cronus', message='Dados do projeto ' + str(project.get()) + ' não encontrados. Por favor verifique se os dados foram salvos corretamente.')
        p1.stop()

def root_destroy():
    root.destroy()

def root_update():
    root.update()
    root.focus_force()

def mainWin():
    splash_win.destroy()
    # root window
    root = tk.Tk()
    root.geometry('300x600')
    root.resizable()
    root.title('Cronus - Carga ao GPE')
    #the GUI layout
    user_label = ttk.Label(root, text="Usuário:").grid(column=0, row=0, sticky=tk.E ,padx=5, pady=5)
    user_entry = ttk.Entry(root, textvariable=user).grid(column=1, row=0, sticky=tk.W ,padx=5, pady=5)

    password_label = ttk.Label(root, text="Senha:").grid(column=0, row=1, sticky=tk.E ,padx=5, pady=5)
    password_entry = ttk.Entry(root, textvariable=password,show='*').grid(column=1, row=1, sticky=tk.W ,padx=5, pady=5)

    project_label = ttk.Label(root, text="Projeto:").grid(column=0, row=2, sticky=tk.E ,padx=5, pady=5)
    project_entry = ttk.Entry(root, textvariable=project).grid(column=1, row=2, sticky=tk.W ,padx=5, pady=5)

    p1 = ttk.Progressbar(root, length=200, mode ="indeterminate", maximum=200,orient=tk.HORIZONTAL).grid(row=3,column=1)
    load_button = ttk.Button(root, text="Carregar", command=load_clicked).grid(column=0, row=3, padx=5, pady=5)
    close_button = ttk.Button(root, text="Fechar", command=root_destroy).grid(column=0, row=22, padx=5, pady=5)
    status=StringVar()
    stauts_label = ttk.Label(root, textvariable=status).grid(column=0, row=23, sticky=tk.W ,padx=5, pady=5, columnspan=2)

splash_win.after(5000, mainWin)

splash_win.mainloop()